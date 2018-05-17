import openpyxl
import datetime
import calendar
from openpyxl.styles import PatternFill, Alignment
import os


def tcgen(sel_month, name, f_name, wnum):

    nextmonth = int(datetime.date.today().month + 1)
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    SHIFTPATH = os.path.join(
        BASE_DIR,
        f'djsg/media/djsg/shifts/{sel_month}のShift.xlsx')
    TCPATH = os.path.join(
        BASE_DIR,
        'djsg/media/djsg/tc.xlsx')

    dates = []
    datesDic = {0: "月", 1: "火", 2: "水", 3: "木", 4: "金", 5: "土", 6: "日"}
    daysinthemonth = []                                     # 月の日にちのリスト

    tc = openpyxl.load_workbook(TCPATH)
    st = openpyxl.load_workbook(SHIFTPATH)

    shift = st[f"{sel_month}月"]
    tt = tc["201404"]
    colA = []
    colB = []
    colC = []
    colD = []
    colE = []
    colF = []
    colG = []
    colH = []
    colI = []
    colJ = []
    colK = []
    colL = []
    colM = []
    colN = []
    colO = []
    colP = []
    colQ = []
    colR = []
    colS = []
    biglist = [colA, colB, colC, colD, colE, colF,
               colG, colH, colI, colJ, colK, colL,
               colM, colN, colO, colP, colQ, colR, colS]
    lis = []
    (a, b) = calendar.monthrange(2018, nextmonth)       # 月の最初の曜日と日数を取得

    for day in range(1, b+1):
        daysinthemonth.append(day)                          # 月の日にちをリストに突っ込む

    thismonth = calendar.Calendar(firstweekday=a)
    dateslist = list(thismonth.iterweekdays()) * 5

    for datenum in dateslist:
        if len(dates) <= b-1:
            dates.append(datesDic[datenum])

    if (0 <= a) and (a <= 5):                               # 月の最初の曜日でパターン分け
        sat = 6 - a
        sun = sat + 1

    elif a == 6:
        sat = 7
        sun = 1
    satlist = []
    sunlist = []
    for i in range(0, 6):                                   # 土曜と日祝をリストに突っ込む
        if (sat + 7 * i) <= b and (sat + 7 * i) not in satlist:
            satlist.append(sat + 7 * i)
        if (sun + 7 * i) <= b:
            sunlist.append(sun + 7 * i)

    for c in shift['A']:
        if c.value == name:
            p_row = c.row
    wsche = []
    for c in shift[p_row]:
        wsche.append(c.value)
    wsche.pop(0)

    for d in range(len(daysinthemonth)):
        colA.append(f"{daysinthemonth[d]}({dates[d]})")

    for v in wsche:

        if v == '×':
            colB.append('H0 公休取得')
            colC.append('H0 公休取得')
            colD.append('')
            colE.append('')
            colF.append('')
            colG.append('')
            colH.append('')
            colI.append('')
            colJ.append('')
            colK.append('')
            colL.append('')
            colM.append('')
            colN.append('')
            colO.append('')
            colP.append('')
            colQ.append('')
            colR.append('')
            colS.append('確定')

        elif v == 1:
            colB.append('契約ス法・出勤')
            colC.append('契約ス法・出勤')
            colD.append('8:30')
            colE.append('16:30')
            colF.append('')
            colG.append('')
            colH.append('14:00')
            colI.append('15:00')
            colJ.append('')
            colK.append('')
            colL.append('')
            colM.append('')
            colN.append('')
            colO.append('')
            colP.append('〇')
            colQ.append('〇')
            colR.append('')
            colS.append('確定')

        elif v == 2:
            colB.append('契約ス法・出勤')
            colC.append('契約ス法・出勤')
            colD.append('8:30')
            colE.append('16:30')
            colF.append('')
            colG.append('')
            colH.append('10:00')
            colI.append('11:00')
            colJ.append('')
            colK.append('')
            colL.append('')
            colM.append('')
            colN.append('')
            colO.append('')
            colP.append('〇')
            colQ.append('〇')
            colR.append('')
            colS.append('確定')
        else:
            colB.append('契約ス法・出勤')
            colC.append('契約ス法・出勤')
            colD.append('10:30')
            colE.append('16:30')
            colF.append('')
            colG.append('')
            colH.append('12:00')
            colI.append('13:00')
            colJ.append('')
            colK.append('')
            colL.append('')
            colM.append('')
            colN.append('')
            colO.append('')
            colP.append('〇')
            colQ.append('〇')
            colR.append('')
            colS.append('確定')
    tt.cell(column=2, row=2, value=wnum)
    fullname = name + " " + f_name
    tt.cell(column=2, row=3, value=fullname)

    for d in range(len(daysinthemonth)):
        l1 = []
        for i in biglist:

            l1.append(i[d])
        lis.append(l1)
        tt.cell(column=2, row=1, value=f"2018年{nextmonth}月")

    for i in range(len(lis)):
        for j in range(len(lis[i])):
            tt.cell(column=j+1, row=i+7, value=lis[i][j])

    satcell = PatternFill(
        patternType='solid',
        start_color='58ACFA',
        end_color='58ACFA')
    suncell = PatternFill(
        patternType='solid',
        start_color='F7819F',
        end_color='F7819F')
    bg = PatternFill(
        patternType='solid',
        start_color='DF0174',
        end_color='DF0174'
    )
    for i in range(len(daysinthemonth)):
        tt.cell(column=4, row=i+7).alignment = Alignment(horizontal='right')
        tt.cell(column=5, row=i+7).alignment = Alignment(horizontal='right')
        tt.cell(column=8, row=i+7).alignment = Alignment(horizontal='right')
        tt.cell(column=9, row=i+7).alignment = Alignment(horizontal='right')
        if tt.cell(column=4, row=i+7).value != '':
            tt.cell(column=4, row=i+7).fill = bg
        if tt.cell(column=5, row=i+7).value != '':
            tt.cell(column=5, row=i+7).fill = bg
        if i in satlist:
            tt.cell(column=1, row=i+6).fill = satcell
        elif i in sunlist:
            tt.cell(column=1, row=i+6).fill = suncell
    tc.save(os.path.join(
        BASE_DIR,
        f"djsg/media/djsg/timecards/{nextmonth}月勤務報告書（{name}）.xlsx"))
