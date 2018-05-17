
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os
from . import month_info

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def shiftgenerator(workerlist):

    workers = list(workerlist.keys())
    workabledays = {}                                       # 従業員が勤務可能な日
    workingdays = {}
    partworkers = []
    fullworkers = []
    noWorkingDays = {4: 8, 5: 6, 6: 10, 7: 8, 8: 12,
                     9: 9, 10: 8, 11: 11, 12: 9}  # 休刊日のディクショナリ
    shift = {}
    
    daysinthemonth = month_info.month_info()["daysinthemonth"]
    satlist = month_info.month_info()["satlist"]
    holydaylist = month_info.month_info()["holydaylist"]
    weekdays = month_info.month_info()["weekdays"]
    nextmonth = month_info.month_info()["nextmonth"]
    dates = month_info.month_info()["dates"]
    a = month_info.month_info()["a"]
    b = month_info.month_info()["b"]

    # 従業員に関する情報の取得

    for worker in workers:
        dayofflist = workerlist[worker]["休み希望"].split(',')
        if workerlist[worker]["属性"] == "土日":
            dayofflist = list(set(weekdays) | set(dayofflist))
            partworkers.append(worker)
        else:
            fullworkers.append(worker)
        try:
            dayofflist = list(map(int, dayofflist))
        except ValueError:
            pass

        workabledays[worker] = list(set(daysinthemonth) - set(dayofflist))
        if noWorkingDays[nextmonth] in workabledays[worker]:
            workabledays[worker].remove(noWorkingDays[nextmonth])

        try:
            workabledays[worker] = list(map(int, workabledays[worker]))
        except:
            pass

    daysforexcel = [f"{nextmonth}月"] + daysinthemonth

    # exelシート作成
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = f"{nextmonth}月"

    wb.create_sheet(index=1, title="workerlist")

    sh = wb["workerlist"]
    sh.merge_cells('A1:C1')
    sh['A1'].font = Font(name="sans-serif", size=15, bold=True)
    sh['A1'] = ("勤務可能日")

    sh.append(daysforexcel)

    # 土日の背景色設定
    satcell = PatternFill(
        patternType='solid',
        start_color='ff00ff00',
        end_color='ff0000ff')
    suncell = PatternFill(
        patternType='solid',
        start_color='ffff0000',
        end_color='ffff0000')

    for i in range(2, b+2):
        selected = sh.cell(row=2, column=i)
        if selected.value in satlist:
            selected.fill = satcell
        elif selected.value in holydaylist:
            selected.fill = suncell

    # 従業員の勤務可能日をシートに反映
    for j in range(len(workers)):
        tictac = []
        for days in daysinthemonth:
            if days in workabledays[workers[j]]:
                tictac.append("○")
            else:
                tictac.append("×")
        workableexcel = [workers[j]] + tictac
        sh.append(workableexcel)
    wb.save(os.path.join(
        BASE_DIR, f'djsg/media/djsg/shifts/{nextmonth}月のShift.xlsx'))

    checklist = weekdays + weekdays + satlist + satlist + holydaylist
    checklist.sort()

    # 従業員の出勤日を確定
    for worker in sh['A']:
        memo = []
        if worker.value in partworkers:     # 土日要員の出勤日
            for sat in satlist:
                if sh.cell(row=worker.row, column=sat+1).value == "○":
                    memo.append(sat)
                    checklist.remove(sat)
            count = 0
            for hol in holydaylist:

                try:
                    if (sh.cell(
                        row=worker.row, column=hol+1).value == "○") and (
                        sh.cell(
                          row=worker.row+1, column=hol+1).value != "○") and (
                          hol in checklist):
                        memo.append(hol)
                        checklist.remove(hol)
                    elif (sh.cell(
                        row=worker.row, column=hol+1).value == "○") and (
                          sh.cell(
                            row=worker.row+1, column=hol+1).value == "○"):
                        if (holydaylist[count-1] not in memo) and (
                              hol in checklist):
                            memo.append(hol)

                            checklist.remove(hol)
                except IndexError:
                    if sh.cell(row=worker.row, column=hol+1).value == "○" and (
                          hol in checklist):
                        memo.append(hol)
                        checklist.remove(hol)
                count += 1
        elif worker.value in fullworkers:   # フルタイムの勤務日
            for day in weekdays:
                if sh.cell(row=worker.row, column=day+1).value == "○":
                    memo.append(day)
                    checklist.remove(day)

        workingdays[worker.value] = memo
    workingdays.pop("勤務可能日")
    workingdays.pop(f"{nextmonth}月")
    try:
        for i in holydaylist:
            for j in partworkers:
                if i in checklist:
                    if i in workabledays[j]:
                        checklist.remove(i)
                        workingdays[j].append(i)

    except ValueError:
        pass

    looptime = 1
    while len(checklist) >= 1:
        try:
            for days in checklist:
                num = 0
                p = ""
                for full in fullworkers:
                    if days in workabledays[full]:
                        if num == 0:
                            num = len(workingdays[full])
                            p = full

                        elif num > len(workingdays[full]):
                            num = len(workingdays[full])
                            p = full

                        if days in workingdays[p]:
                            if p == fullworkers[0]:
                                p = fullworkers[1]
                            else:
                                p = fullworkers[0]

                        checklist.remove(days)
                        workingdays[p].append(days)

        except ValueError:
            pass
        
        looptime += 1
        
        if looptime >= 40:
            break

    p = ""
    num = 0
    for l in range(0, len(partworkers)):

        if num == 0:
            num = len(workingdays[partworkers[l]])
            p = partworkers[l]
        elif num <= len(workingdays[partworkers[l]]):
            num = len(workingdays[partworkers[l]])  # 長い方を入れている
            p = partworkers[l]

    for l in range(0, len(partworkers)):
        if len(workingdays[partworkers[l]]) == num:
            pass
        elif num - len(workingdays[partworkers[l]]) >= 2:

                longerset = set(workingdays[p])
                fixingset = longerset & set(holydaylist)
                fixingp = partworkers[l]
                for d in fixingset:
                    if num - len(workingdays[partworkers[l]]) >= 2:
                        if d not in workingdays[fixingp]:
                            if d in workabledays[fixingp]:
                                workingdays[p].remove(d)
                                workingdays[fixingp].append(d)

    p = ""
    num = 0
    for l in range(0, len(fullworkers)):

        if num == 0:
            num = len(workingdays[fullworkers[l]])
            p = fullworkers[l]
        elif num <= len(workingdays[fullworkers[l]]):
            num = len(workingdays[fullworkers[l]])  # pに長い方を入れている
            p = fullworkers[l]
            # pは長く働く人の名前、numは長く働く人の勤務日数
    for l in range(0, len(fullworkers)):
        if len(workingdays[fullworkers[l]]) == num:
            pass
        elif num - len(workingdays[fullworkers[l]]) >= 2:

            longerset = set(workingdays[p])
            fixingset = longerset & set(satlist)
            fixingp = fullworkers[l]

            for d in fixingset:
                if num - len(workingdays[fullworkers[l]]) >= 2:
                    if d not in workingdays[fixingp]:
                        if d in workabledays[fixingp]:
                            workingdays[p].remove(d)
                            workingdays[fixingp].append(d)

    wb.create_sheet(index=2, title="workingday")
    sh2 = wb["workingday"]
    sh2.append(['出勤日'])
    sh2.append(daysforexcel)
    for j in range(len(workers)):
        tictac = []
        for days in daysinthemonth:
            if days in workingdays[workers[j]]:
                tictac.append("○")
            else:
                tictac.append("×")
        workingexcel = [workers[j]] + tictac
        sh2.append(workingexcel)

    satcell = PatternFill(
        patternType='solid',
        start_color='ff00ff00',
        end_color='ff0000ff')
    suncell = PatternFill(
        patternType='solid',
        start_color='ffff0000',
        end_color='ffff0000')

    for i in range(2, b+2):
        selected = sh2.cell(row=2, column=i)
        selected.alignment = Alignment(horizontal='center')
        selected.font = Font(size=15, bold=True)
        if selected.value in satlist:
            selected.fill = satcell
        elif selected.value in holydaylist:
            selected.fill = suncell

    wb.save(
        os.path.join(BASE_DIR,
                     f'djsg/media/djsg/shifts/{nextmonth}月のShift.xlsx'))

    hardworker = ""
    wdays = 0
    for h in workers:
        if wdays == 0:
            wdays = len(workabledays[h])
            hardworker = h

        elif len(workabledays[h]) >= wdays:
            wdays = len(workabledays[h])
            hardworker = h

    sheet.append(['日勤シフト']+dates)
    sheet.append(daysforexcel)

    for i in range(1, b+2):
        selected = sheet.cell(row=2, column=i)
        selected.alignment = Alignment(horizontal='center')
        selected.font = Font(size=15, bold=True)
        if selected.value in satlist:
            selected.fill = satcell
        elif selected.value in holydaylist:
            selected.fill = suncell

    for worker in sh2['A']:
        counter = 1
        shiftlist = []
        if worker.value == hardworker:
            row = worker.row
            for day in sh2[row]:
                if day.value == "○":
                    thiscolumn = day.column.upper()
                    valuelist = []
                    marker = "A"
                    for cel in sheet[str(thiscolumn)]:
                        valuelist.append(cel.value)
                    if valuelist[1] in holydaylist:
                        shiftlist.append("出")
                        marker = "B"
                    valuelist.pop(1)
                    if marker == "A":
                        shiftlist.append(counter)
                    if counter == 1:
                        counter = 2
                    else:
                        counter = 1
                elif day.value == "×":
                    shiftlist.append("×")
            shiftlist = [hardworker] + shiftlist
            sheet.append(shiftlist)
            shift[worker.value] = shiftlist
            wb.save(os.path.join(
                BASE_DIR,
                f'djsg/media/djsg/shifts/{nextmonth}月のShift.xlsx'))

    for worker in sh2['A']:
        shiftlist = []
        if worker.value != hardworker:
            if worker.value in workers:
                row = worker.row
                for day in sh2[row]:
                    if day.value == "○":
                        thiscolumn = day.column.upper()
                        valuelist = []
                        marker = "A"
                        for cel in sheet[str(thiscolumn)]:
                            valuelist.append(cel.value)
                        if valuelist[1] in holydaylist:
                            shiftlist.append("出")
                            marker = "B"
                        valuelist.pop(1)
                        if marker == "A":
                            if 1 in valuelist:
                                shiftlist.append(2)
                            elif 2 in valuelist:
                                shiftlist.append(1)
                            else:
                                shiftlist.append(1)
                    elif day.value == "×":
                        shiftlist.append("×")

                shiftlist = [worker.value] + shiftlist
                shift[worker.value] = shiftlist
                sheet.append(shiftlist)
                wb.save(os.path.join(
                    BASE_DIR,
                    f'djsg/media/djsg/shifts/{nextmonth}月のShift.xlsx'))

    for row in sheet.rows:
        for cell in row:
            if cell.value:
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    outline=True,
                    left=Side(style="medium", color="FF000000"),
                    right=Side(style="medium", color="FF000000"),
                    top=Side(style="medium", color="FF000000"),
                    bottom=Side(style="medium", color="FF000000")
                    )
    noworkingcell = PatternFill(
        patternType='solid',
        start_color='ff615a5a',
        end_color='ff615a5a')

    for days in sheet['2']:
        if days.value == noWorkingDays[nextmonth]:
            dcolumn = days.column  # 縦がcolumn 横がrow
            for selected in sheet[dcolumn]:
                selected.fill = noworkingcell
    wb.save(os.path.join(
        BASE_DIR,
        f'djsg/media/djsg/shifts/{nextmonth}月のShift.xlsx'))
